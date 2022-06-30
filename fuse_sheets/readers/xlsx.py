from copy import copy
from io import BytesIO
from logging import getLogger
from typing import Tuple, Generator, Any, Union, Awaitable

from openpyxl.cell import ReadOnlyCell
from openpyxl.cell.read_only import EmptyCell
from openpyxl.reader import excel

from .base import ISheetReader
from fuse_core.handlers.containers import FuseDictionary


logger = getLogger(__name__)


class XlsxSheetReader(ISheetReader):

    @staticmethod
    def read_file(content: Union[str, bytes]) -> Any:
        """
        Read file and, for example, return `openpyxl.Workbook`
        """
        if hasattr(getattr(content, 'file', None), 'getvalue'):
            content = BytesIO(getattr(content.file, 'getvalue', None)())

        return excel.load_workbook(
            filename=content,
            keep_vba=False,
            read_only=True,
            data_only=True
        )

    @staticmethod
    def find_sheet(sheet_data, verbose_names: Tuple[str, ...]) -> Any:
        """
        Find needed sheet
        """
        for sheet in sheet_data:
            for data in sheet.iter_rows(min_row=1, max_row=1):
                if tuple(i.value for i in data if i.value) == verbose_names:
                    return sheet

    @staticmethod
    def get_fuse_dictionary(
        sheet_data: Any,
        headers: Tuple["Field", ...]
    ) -> Union[Awaitable[Generator], Generator]:
        """
        Convert any data to `FuseDictionary` and yield it
        """
        for col in sheet_data.iter_rows(2, max_col=len(headers)):
            fuse_dict = FuseDictionary()

            for ri, row in enumerate(col):
                field_inst = copy(headers[ri])
                fuse_dict[field_inst.name] = field_inst
                try:
                    if isinstance(row, (ReadOnlyCell, EmptyCell)):
                        field_inst.validate(row.value)
                        field_inst.set(row.value)

                except Exception as e:
                    logger.critical(e)

                del row

            yield fuse_dict
