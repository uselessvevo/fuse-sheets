from copy import copy
from io import BytesIO
from typing import Tuple, Generator, Any, Union, Awaitable

from openpyxl.cell import ReadOnlyCell, Cell
from openpyxl.cell.read_only import EmptyCell
from openpyxl.reader import excel

from fuse_sheets.core.readers.abc import ISheetReader
from fuse_core.core.containers import FieldDictionary
from fuse_core.core.exceptions import ValueValidationError


class XlsxSheetReader(ISheetReader):

    def read_file(self, content: Union[str, bytes]) -> Any:
        """
        Read file and, for example, return `openpyxl.Workbook`
        """
        if hasattr(getattr(content, 'file', None), 'getvalue'):
            content = BytesIO(getattr(content.file, 'getvalue', None)())

        return excel.load_workbook(
            filename=content,
            keep_vba=False,
            read_only=True,
            data_only=True
        )

    def get_max_row(self, workbook) -> int:
        return workbook.max_row

    def find_sheet(
        self,
        sheet_data: bytes,
        verbose_names: Tuple[str, ...]
    ) -> Any:
        for sheet in sheet_data:
            for data in sheet.iter_rows(min_row=1, max_row=1):
                if tuple(i.value for i in data if i.value) == verbose_names:
                    return sheet

    def get_fuse_dictionary(
        self,
        sheet_data: Any,
        headers: Tuple["Field", ...]
    ) -> Union[Awaitable[Generator], Generator]:
        """
        Convert any data to `FuseDictionary` and yield it
        """
        for col in sheet_data.iter_rows(2, max_col=len(headers)):
            fuse_dict = FieldDictionary()

            for ri, row in enumerate(col):
                field_inst = copy(headers[ri])
                fuse_dict[field_inst.name] = field_inst
                try:
                    if isinstance(row, (Cell, ReadOnlyCell, EmptyCell)):
                        field_inst.validate(row.value)
                        field_inst.set(row.value)

                except ValueValidationError:
                    self._logger.warning(f'Can\'t validate value {row.value}')

                except Exception:
                    self._logger.warning('Something went wrong')

                del row

            yield fuse_dict
