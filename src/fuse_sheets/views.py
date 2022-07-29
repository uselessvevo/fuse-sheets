"""
Base mapper class
"""
from io import BytesIO
from typing import Tuple

from uuid import uuid4
from cgi import FieldStorage

import openpyxl
from openpyxl.writer.excel import save_virtual_workbook

from fuse_core.core.fields import Field

from aiohttp import web
from aioify import aioify


class FuseSheetsView(web.View):

    # Sheet uploader version
    version: int = 1

    # URL pattern to our mapper endpoint
    urlpattern: str

    # List of `Field` objects
    headers: Tuple[Field]

    task_class: "FuseSheetsTask"

    @aioify
    def _write_headers(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet(index=0)
        headers = [i.verbose_name for i in self.headers]

        for col, value in enumerate(headers):
            worksheet.cell(1, col + 1).value = value

        return save_virtual_workbook(workbook)

    async def get(self) -> web.Response:
        """
        Get method returns template
        """
        workbook = await self._write_headers()
        response = web.Response(body=workbook, status=201)
        response.headers['Content-type'] = 'application/vnd.ms-excel'
        response.headers['Content-Disposition'] = f'attachment; filename={self.urlpattern}_{str(uuid4())[:4]}.xlsx'
        await response.prepare(self.request)
        return response

    async def post(self) -> web.Response:
        data = await self.request.post()
        file_data = FieldStorage(BytesIO(data.get('file').file.read()))
        task_class = await self.task_class(self.headers)
        await task_class.prepare(file_data, data.get('file').filename)
        return web.Response(status=201)
